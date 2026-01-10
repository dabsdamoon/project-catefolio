from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class TemplateService:
    def __init__(self, template_path: Path) -> None:
        self.template_path = template_path

    def build_template_bytes(self, transactions: list[dict[str, Any]]) -> bytes:
        wb = load_workbook(self.template_path)
        ws = wb.active

        date_row = self._find_row(ws, "날짜")
        credit_row = self._find_row(ws, "입금")
        debit_row = self._find_row(ws, "지출")
        remark_row = self._find_row(ws, "비고")
        if date_row is None or credit_row is None or debit_row is None:
            raise ValueError("Template missing required labels.")
        if remark_row is None:
            remark_row = ws.max_row + 1

        date_columns = self._extract_date_columns(ws, date_row)
        if not date_columns:
            raise ValueError("Template has no date columns.")

        debit_start = self._find_first_data_row(ws, debit_row + 1, remark_row, date_columns)
        credit_rows = list(range(credit_row, debit_row))
        debit_rows = list(range(debit_start, remark_row))

        self._clear_section(ws, credit_rows, date_columns)
        self._clear_section(ws, debit_rows, date_columns)

        grouped = self._group_transactions(transactions)
        for date_key, cols in date_columns.items():
            credits = grouped.get(date_key, {}).get("credit", [])
            debits = grouped.get(date_key, {}).get("debit", [])
            self._write_entries(ws, credit_rows, cols, credits)
            self._write_entries(ws, debit_rows, cols, debits)

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def _find_row(ws, label: str) -> int | None:
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, 1).value == label:
                return row
        return None

    @staticmethod
    def _extract_date_columns(ws, date_row: int) -> dict[str, tuple[int, int]]:
        date_columns: dict[str, tuple[int, int]] = {}
        for col in range(2, ws.max_column + 1, 2):
            value = ws.cell(date_row, col).value
            if not value:
                continue
            if isinstance(value, datetime):
                date_key = value.date().isoformat()
            else:
                date_key = str(value).strip()
            date_columns[date_key] = (col, col + 1)
        return date_columns

    @staticmethod
    def _find_first_data_row(ws, start: int, end: int, date_columns: dict[str, tuple[int, int]]) -> int:
        for row in range(start, end):
            for col, amt_col in date_columns.values():
                if ws.cell(row, col).value or ws.cell(row, amt_col).value:
                    return row
        return start

    @staticmethod
    def _clear_section(ws, rows: list[int], date_columns: dict[str, tuple[int, int]]) -> None:
        for row in rows:
            for col, amt_col in date_columns.values():
                ws.cell(row, col).value = None
                ws.cell(row, amt_col).value = None

    @staticmethod
    def _group_transactions(transactions: list[dict[str, Any]]) -> dict[str, dict[str, list[dict[str, Any]]]]:
        grouped: dict[str, dict[str, list[dict[str, Any]]]] = {}
        for tx in transactions:
            date_key = tx.get("date")
            if not date_key:
                continue
            bucket = grouped.setdefault(date_key, {"credit": [], "debit": []})
            if tx.get("amount", 0) >= 0:
                bucket["credit"].append(tx)
            else:
                bucket["debit"].append(tx)
        return grouped

    @staticmethod
    def _write_entries(
        ws,
        rows: list[int],
        cols: tuple[int, int],
        entries: list[dict[str, Any]],
    ) -> None:
        desc_col, amt_col = cols
        for idx, entry in enumerate(entries):
            if idx >= len(rows):
                break
            row = rows[idx]
            ws.cell(row, desc_col).value = entry.get("description", "")
            ws.cell(row, amt_col).value = abs(entry.get("amount", 0))

